#!/usr/bin/python
# -*- coding:utf-8 -*-
# FileName:     rule_retrieve.py
# Description:  excel表规则检索,寻找出指定范围内符合规则的关键字
# Author:       Wangjia
# Version:      0.0.1
# Created:      2019-05-24
# Company:      None
# LastChange:   create 2019-05-24
# History:      
#----------------------------------------------------------------------------

# 规则
# 列名 “信号”
# 行初始 2
# 行结束 last
# 寻找关键字 "绿灯"

MAX=65536

import openpyxl
from xml_config_parse import xml_config_parse
import os
import re

class excel_rule():
    s_inputFileName=""
    s_wb=''
    s_trainDir="UP"
    s_stationList=[] #车站名列表
    s_stationWithRowDicList=[]
    s_maxRow=0
    def __init__(self,inputFileName="-Unlicensed-18088-5027.05150935f.xlsx"):
        self.s_inputFileName=inputFileName
        os.chdir("../input")
        self.s_wb = openpyxl.load_workbook(self.s_inputFileName)
        os.chdir("../src")       
        self.ElRu_CaluteDir()
        self.s_stationList=self.ElRu_GetStationList()
        self.s_stationWithRowDicList=self.ElRu_GetStationWithRowNum()
        self.s_maxRow=self.ElRu_GetMaxRowNum()
        pass

        
    # 获取表格中内容
    def ElRu_GetWbCont(self):
        sheets = self.s_wb.get_sheet_names()
        
        sheet=self.s_wb.get_sheet_by_name(sheets[0])
        for r in range(1,sheet.max_row+1):
            if r == 1:
                for c in range(1,sheet.max_column+1):#
                    b=sheet.cell(row=r,column=c).value
                    #print('\n'+''.join([str(sheet.cell(row=r,column=c).value).ljust(17) for c in range(1,sheet.max_column+1)] ))
                    #print(type(sheet.cell(row=r,column=c).value))
            else:
                for c in range(1,sheet.max_column+1):
                    #print(''.join([str(sheet.cell(row=r,column=c).value).ljust(20) ] ))
                    a=sheet.cell(row=r,column=c).value
    
    # 根据规则获取关键字
    def ElRu_GetKey(self,firstRow=0,lastRow=0, col=0, key=u"",outCol=0, frontToRear=True, fullPattern=True):
        sheet=self.s_wb.worksheets[0]
        #步长
        step=1
        # 从最后一行开始找,最大参数65536表示
        if lastRow==MAX:
            lastRow=sheet.max_row
        # 直接根据固定的行列寻找关键字
        if frontToRear==False:
            temp=firstRow
            firstRow=lastRow
            lastRow=temp
            step=-1

            
        if fullPattern==True:
            i=firstRow
            for i in range(firstRow,lastRow+1,step):
                if(sheet.cell(row=i,column=col).value==(key)):
                    return sheet.cell(row=i,column=outCol).value
        else:
            i=firstRow
            for i in range(firstRow,lastRow+1,step):
                #部分匹配 比如说车站 神池南---可能为 神池南I场 or 神池南II场
                if(re.search(key,sheet.cell(row=i,column=col).value)):
                    return sheet.cell(row=i,column=outCol).value
        return "not found"   
    
    #获取关键字行数
    def ElRu_GetRowNum(self,firstRow=0,lastRow=0, col=0, key=u"",outCol=0, frontToRear=True, fullPattern=True):
        _dic={}
        sheet=self.s_wb.worksheets[0]
        #步长
        step=1
        # 从最后一行开始找,最大参数65536表示
        if lastRow==MAX:
            lastRow=sheet.max_row
        # 直接根据固定的行列寻找关键字
        if frontToRear==False:
            temp=firstRow
            firstRow=lastRow
            lastRow=temp
            step=-1
            
        if fullPattern==True:
            i=firstRow
            for i in range(firstRow,lastRow+1,step):
                if(sheet.cell(row=i,column=col).value==(key)):
                    _dic[key]=i
                    return _dic
        else:
            i=firstRow
            for i in range(firstRow,lastRow+1,step):
                #部分匹配 比如说车站 神池南---可能为 神池南I场 or 神池南II场
                if(re.match(key,str(sheet.cell(row=i,column=col).value))):
                    _dic[key]=i
                    return _dic

    # 获取最大行数   
    def ElRu_GetMaxRowNum(self):
        sheet=self.s_wb.worksheets[0]  
        return sheet.max_row         
         
    # 返回数据参考文件
    def ElRu_GetRefFileName(self):
        return self.s_inputFileName

    #车次
    def ElRu_CaluteDir(self):
        num=self.ElRu_GetKey(5,5,2,key=u"车次",outCol=5)
        num=int(num)%2
        if num == 1:
            self.s_trainDir=u"下行"
        else:
            self.s_trainDir=u"上行" 
    
    def ElRu_GetDir(self):
        return self.s_trainDir

    # 返回车站列表
    def ElRu_GetStationList(self):
        parseObj=xml_config_parse()
        _stationList=parseObj.get_station_list()   
        if(self.s_trainDir==u"下行"):
            _stationList.reverse()
            return _stationList
        return _stationList
    
    # 返回车站名对应的行号 字典列表
    def ElRu_GetStationWithRowNum(self):
        stationRowNumDicList=[]
        g_stationList=self.s_stationList
        for station in g_stationList:
            _dic=self.ElRu_GetRowNum(1,MAX,5,key=station,outCol=0, frontToRear=False,fullPattern=False)
            stationRowNumDicList.append(_dic) 
            #stationRowNumDicList.reverse()
        return stationRowNumDicList
    
    # 返回起始站
    def ElRu_GetStartStationName(self):
        for i in range(len(self.s_stationWithRowDicList)):
            if self.s_stationWithRowDicList[i]!=None:
                for key,value in self.s_stationWithRowDicList[i].items():
                    return key
        return "not found start station"

    # 返回终点
    def ElRu_GetEndStationName(self):
        for i in range(len(self.s_stationWithRowDicList)-1,0,-1):
            if self.s_stationWithRowDicList[i]!=None:
                for key,value in self.s_stationWithRowDicList[i].items():
                    return key
        return "not found end station"

    # 返回车站 "侧线股道号"
    def ElRu_GetStationSideTrackNum(self,i):
        '''
                    侧线股道号  "车站" 行号后的1行或者两行都可能存在 侧线号
        '''
        '''
                        第一站是没有侧线概念的,所以肯定是从第二站开始
        '''
        if i==0 and self.s_stationWithRowDicList[0]==None:
            return u"本次未通过该站"
        elif i==0:
            return u"正线进站"
        value=0
        nextValue=0
        _dic=self.s_stationWithRowDicList[i-1]
        if _dic==None:
            return u"本次未通过该站"
        for _key,_value in _dic.items():
            value=_value
        _nextDic=self.s_stationWithRowDicList[i]
        if _nextDic==None:
            return u"本次未通过该站"
        for _key,_value in _nextDic.items():
            nextValue=_value
        
        str1=self.ElRu_GetKey(value,nextValue,2,key=u"侧线选择",outCol=5)
        if(str1=="not found"):
            pass
        else:
            return str1        
        pass       
        
        num1=''
        num2=''  
        value=0      
        sheet=self.s_wb.worksheets[0]  
        _dic=self.s_stationWithRowDicList[i]
        if _dic ==None:
            return u"本次未通过该站"
        for _key,_value in _dic.items():
            value=_value        
        row1=value
        row1=row1+2
        row2=row1+1

        # 第G列的内容,包含进站侧线股道号，从中提取 数字.例如 "进站SL6"
        if sheet.cell(row=row1,column=7).value!=None:
            for c in sheet.cell(row=row1,column=7).value:
                if c>='0' and c<='9':
                    num1=num1+c

        # 第G列的内容,包含进站侧线股道号，从中提取 数字.例如 "进站SL6"
        if sheet.cell(row=row2,column=7).value!=None:
            for c in sheet.cell(row=row2,column=7).value:
                if c>='0' and c<='9':
                    num2=num2+c  
        if  num1!='':
            return num1
        elif num2!='':               
            return num2
        else:
            return u"正线进站"
    
    # 返回车站 进站停车距离
    def ElRu_GetStationStopDistance(self,i=0):
        '''
        param: station从xml表中的第一个station到最后一个
        '''
        value=0
        nextValue=0
        _dic=self.s_stationWithRowDicList[i]
        if _dic==None:
            return u"本次未通过该站"
        for _key,_value in _dic.items():
            value=_value
        if i!=len(self.s_stationWithRowDicList)-1:  
            _nextDic=self.s_stationWithRowDicList[i+1]
            if _nextDic==None:
                nextValue=self.ElRu_GetMaxRowNum()
            else:
                for _key,_value in _nextDic.items():
                    nextValue=_value
        else:
            nextValue=self.s_maxRow
        str1=self.ElRu_GetKey(value,nextValue,2,key=u"停车",outCol=6,frontToRear=True, fullPattern=False)
        if(str1=="not found"):
            return u"正线通过"
        else:
            return str1
        pass

    # 返回车站 机外停车距离
    def ElRu_GetOutStationStopDistance(self,i=0):
        '''
                        第一站是没有机外停车概念的,所以肯定是从第二站开始
        '''
        if i==0:
            return u"无机外停车"
        value=0
        nextValue=0
        _dic=self.s_stationWithRowDicList[i-1]
        if _dic==None:
            return u"本次未通过该站"
        for _key,_value in _dic.items():
            value=_value
        _nextDic=self.s_stationWithRowDicList[i]
        if _nextDic==None:
            return u"本次未通过该站"
        for _key,_value in _nextDic.items():
            nextValue=_value
        
        str1=self.ElRu_GetKey(value,nextValue,2,key=u"机外停车",outCol=6)
        if(str1=="not found"):
            return "无机外停车"
        else:
            return str1        
        pass
    
    # 根据输入excel获取输出行 列表
    def ElRu_GetRowList(self):
        #1日期
        list=[]
        list.append(self.ElRu_GetKey(2,2,2,key=u"文件头",outCol=5))
        #2参考数据文件
        list.append(self.ElRu_GetRefFileName())
        #3机车号
        list.append(self.ElRu_GetKey(15,15,2,key=u"机车号",outCol=5))
        #4车次 
        list.append(self.ElRu_GetKey(5,5,2,key=u"车次",outCol=5))
        #5司机号=司机1+司机2
        list.append(u"司机号1 "+self.ElRu_GetKey(10,10,2,key=u"司机1",outCol=5)+ u" 司机号2 "+self.ElRu_GetKey(11,11,2,key=u"司机2",outCol=5))
        #6总重 
        list.append(self.ElRu_GetKey(12,12,2,key=u"总重",outCol=5))
        #7计长 
        list.append(self.ElRu_GetKey(14,14,2,key=u"计长",outCol=5))
        #8关门车数量 
        list.append(0)
        #9里程
        list.append(abs(float(self.ElRu_GetKey(15,MAX,2,key=u"进站",outCol=4,frontToRear=False))-float(self.ElRu_GetKey(2,MAX,2,key=u"出站",outCol=4,frontToRear=True))))

        #10开车侧线号
        list.append(self.ElRu_GetKey(63,63,2,key=u"开车侧线号",outCol=5))
        #11 起始站
        list.append(self.ElRu_GetStartStationName())
        #12 终点站
        list.append(self.ElRu_GetEndStationName())   
             
        for i in range(0,len(self.ElRu_GetStationList())):
            # 侧线股道号 --> 进站道岔 搜寻进站下一行的7列 例如 "进站SL6",表示侧线股道号6
            list.append(self.ElRu_GetStationSideTrackNum(i))
            # 进站停车距离 --> 搜寻 两站之间的 “停车”
            list.append(self.ElRu_GetStationStopDistance(i))
            # 机外停车距离
            list.append(self.ElRu_GetOutStationStopDistance(i))
        return list
    
if __name__ == '__main__':           
    rule=excel_rule()  
    g_stationList=rule.ElRu_GetStationList()   
    g_stationDicList=rule.ElRu_GetStationWithRowNum() 
    
    print(g_stationDicList)
    
    print(g_stationList)
    
    #1日期   
    print(rule.ElRu_GetKey(2,2,2,key=u"文件头",outCol=5))
    #2机车号
    print(rule.ElRu_GetKey(15,15,2,key=u"机车号",outCol=5))
    #3参考数据文件 
    print(rule.ElRu_GetRefFileName())
    #4车次 
    print(rule.ElRu_GetKey(5,5,2,key=u"车次",outCol=5))
    #5司机号=司机1+司机2
    print(rule.ElRu_GetKey(10,10,2,key=u"司机1",outCol=5)+ rule.ElRu_GetKey(11,11,2,key=u"司机2",outCol=5))
    #6总重 
    print(rule.ElRu_GetKey(12,12,2,key=u"总重",outCol=5))
    #7计长 
    print(rule.ElRu_GetKey(14,14,2,key=u"计长",outCol=5))
    #8关门车数量 
    print(0)
    #9里程
    print(abs(float(rule.ElRu_GetKey(15,MAX,2,key=u"进站",outCol=4,frontToRear=False))-float(rule.ElRu_GetKey(2,MAX,2,key=u"出站",outCol=4,frontToRear=True))))
    
    #行坐标  从后往前搜索
    
    # 侧线股道号 --> 进站道岔 搜寻进站下一行的7列 例如 "进站SL6",表示侧线股道号6
    
    for i in range(0,len(g_stationDicList)):
        print(rule.ElRu_GetStationSideTrackNum(i))
        print(rule.ElRu_GetStationStopDistance(i))
        print(rule.ElRu_GetOutStationStopDistance(i))
    
    
    # 进站停车距离 -->  两站间搜寻 "停车",第一个为有效值
    print(rule.ElRu_GetRowList())
