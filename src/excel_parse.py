#!/usr/bin/python
# -*- coding:utf-8 -*-
# FileName:     excel_parse.py
# Description:  
# Author:       Wangjia
# Version:      0.0.1
# Created:      2019-05-14
# Company:      None
# LastChange:   create 2019-05-14
# History:      
#----------------------------------------------------------------------------
from xml_config_parse import xml_config_parse

'''
建立一个类，根据关键字-规则，得到内容
'''
import openpyxl 
import os

class excel_parse():
    s_inputDir=""
    s_inputFileName=""
    s_inputFileIndex=-1
    s_outputFileName=""
    
    def __init__(self,inputDir="",outputFileName="outputFileName.xlsx"):
        self.s_inputDir=inputDir
        self.s_outputFileName=outputFileName
        pass
    
    # "信号"列无绿灯,则该文件无效,判断15数据是否是在自动驾驶下
    def IsExcelContVaild(self):
        _inExcelName=self.GetIterFileName()
        
    # 根据上下行来决定将 List添加到哪个sheet
    # 如果excel不存在那么创建
    def AddExcelRow(self, RowList=[], Dir=u"上行"):
        
        if Dir==u"上行":
            oppsiteDir=u"下行"
        elif Dir==u"下行":
            oppsiteDir=u"上行"
            
        try:
            wb=openpyxl.load_workbook(self.s_outputFileName)
            sheetU=wb.get_sheet_by_name(Dir)
        except: 
            wb = openpyxl.Workbook()
            sheetU = wb.create_sheet(Dir, 0)
            sheetL = wb.create_sheet(oppsiteDir)
            #wb.save(self.s_outputFileName)        

        # Data can be assigned directly to cells
        for _dic in RowList:
            #print(type(dir))
            for item,value in _dic.items():
                sheetU[item]=value
                print(item,value)
        # Save the file
        wb.save(self.s_outputFileName)

        pass
    
    # Iter获取文件名 ,每调用一次迭代 test ok
    def GetIterFileName(self):
        self.s_inputFileIndex=self.s_inputFileIndex+1
        #改变当前路径
        os.chdir(self.s_inputDir)
        #获取当前路径文件夹名
        file_dir=os.getcwd()
        #当前文件夹目录，子目录，文件
        for root, dirs, files in os.walk(file_dir): 
            if(self.s_inputFileIndex<len(files)):
                return files[self.s_inputFileIndex]
            else:
                return None
        pass
    

xmlObj=xml_config_parse()
g_InFirstRowDict=xmlObj.get_row_format(xmlObj.get_first_row_list())
print(g_InFirstRowDict)
g_excelObj=excel_parse("../input","outputFileName.xlsx")
g_excelObj.AddExcelRow(g_InFirstRowDict)

